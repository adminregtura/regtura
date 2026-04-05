"""Excel parser for FINREP templates.

Reads an Excel file in the standard EBA FINREP layout and converts it
into a SubmissionData object that the validation engine can process.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

from regtura.common import SubmissionData

CELL_REF_PATTERN = re.compile(r"^r\d{3}c\d{3}$")

SHEET_TO_TEMPLATE = {
    "F 01.01": "F 01.01", "F01.01": "F 01.01", "F0101": "F 01.01",
    "Balance Sheet": "F 01.01",
}


def parse_excel(
    file_path: str | Path | None = None,
    file_obj: BinaryIO | None = None,
    framework: str = "finrep",
    period: str = "unknown",
) -> SubmissionData:
    if file_path is None and file_obj is None:
        raise ValueError("Provide either file_path or file_obj.")

    wb = load_workbook(file_path or file_obj, data_only=True)
    templates: dict[str, dict[str, float]] = {}
    metadata: dict = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        template_id = _match_template(sheet_name)
        sheet_meta = _extract_metadata(ws)
        metadata.update(sheet_meta)
        if sheet_meta.get("period"):
            period = sheet_meta["period"]
        template_data = _extract_template_data(ws)
        if template_data:
            templates[template_id or "F 01.01"] = template_data

    wb.close()

    if not templates:
        raise ValueError(
            "No recognisable FINREP data found. Ensure column A contains "
            "cell references (e.g. r010c010) and column C contains values."
        )

    return SubmissionData(framework=framework, period=period, templates=templates, metadata=metadata)


def _match_template(sheet_name: str) -> str | None:
    clean = sheet_name.strip()
    if clean in SHEET_TO_TEMPLATE:
        return SHEET_TO_TEMPLATE[clean]
    normalised = clean.replace(" ", "").replace(".", "").upper()
    for key, tid in SHEET_TO_TEMPLATE.items():
        if key.replace(" ", "").replace(".", "").upper() == normalised:
            return tid
    return None


def _extract_metadata(ws) -> dict:
    metadata = {}
    for row in ws.iter_rows(min_row=1, max_row=7, max_col=3, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip().lower()
                if "entity" in val or "institution" in val:
                    nxt = ws.cell(row=cell.row, column=cell.column + 1)
                    if nxt.value:
                        metadata["entity"] = str(nxt.value).strip()
                elif "period" in val or "date" in val:
                    nxt = ws.cell(row=cell.row, column=cell.column + 1)
                    if nxt.value:
                        metadata["period"] = str(nxt.value).strip()
                elif "currency" in val:
                    nxt = ws.cell(row=cell.row, column=cell.column + 1)
                    if nxt.value:
                        metadata["currency"] = str(nxt.value).strip()
    return metadata


def _extract_template_data(ws) -> dict[str, float]:
    data = {}
    for row in ws.iter_rows(min_row=1, max_col=10, values_only=False):
        cell_ref = None
        value = None
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                clean = cell.value.strip().lower()
                if CELL_REF_PATTERN.match(clean):
                    cell_ref = clean
                    vc = ws.cell(row=cell.row, column=3)
                    if vc.value is not None and _is_numeric(vc.value):
                        value = float(vc.value)
                    else:
                        for sc in row:
                            if sc.column > cell.column and _is_numeric(sc.value):
                                value = float(sc.value)
                                break
                    break
        if cell_ref and value is not None:
            data[cell_ref] = value
    return data


def _is_numeric(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value.replace(",", ""))
            return True
        except ValueError:
            return False
    return False